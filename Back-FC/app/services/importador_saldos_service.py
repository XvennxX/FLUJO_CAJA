"""
Servicio para importar saldos iniciales desde Excel con múltiples hojas (una por día)
"""
import io
from datetime import date, datetime, timedelta
from typing import Dict, List, Tuple, Optional
from decimal import Decimal
import re

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.transacciones_flujo_caja import TransaccionFlujoCaja, AreaTransaccion
from app.models.conceptos_flujo_caja import ConceptoFlujoCaja
from app.models.cuentas_bancarias import CuentaBancaria
from app.models.trm import TRM


class ImportadorSaldosResult:
    def __init__(self) -> None:
        self.cuentas_procesadas: int = 0
        self.cuentas_sin_match: List[str] = []
        self.detalle_cuentas: List[Dict] = []
        self.dias_procesados: int = 0
        self.dias_sin_trm: List[str] = []
        self.errores: List[str] = []
        self.tipo_carga: str = ""
        self.overwrite: bool = False
        self.mes: str = ""
        self.dia: Optional[str] = None
        self.cuentas_tesoreria: int = 0
        self.cuentas_pagaduria: int = 0

    def to_dict(self):
        return {
            "tipo_carga": self.tipo_carga,
            "mes": self.mes,
            "dia": self.dia,
            "overwrite": self.overwrite,
            "cuentas_procesadas": self.cuentas_procesadas,
            "cuentas_tesoreria": self.cuentas_tesoreria,
            "cuentas_pagaduria": self.cuentas_pagaduria,
            "cuentas_sin_match": self.cuentas_sin_match,
            "detalle_cuentas": self.detalle_cuentas,
            "dias_procesados": self.dias_procesados,
            "dias_sin_trm": self.dias_sin_trm,
            "errores": self.errores,
        }


class ImportadorSaldosService:
    @staticmethod
    def _parse_excel_multi_sheet(xls_bytes: bytes, etiqueta_fila_objetivo: str, mes: str) -> Dict[date, Tuple[Dict[str, Decimal], Dict[str, bool]]]:
        """
        Parsea un Excel con múltiples hojas (una por día).
        Retorna: Dict[fecha -> (Dict[cuenta->valor], Dict[cuenta->es_usd])]
        """
        import logging
        from datetime import datetime
        logger = logging.getLogger(__name__)
        logger.info(f"=== PARSE MULTI-HOJA: Buscando '{etiqueta_fila_objetivo}' ===")
        
        wb = load_workbook(io.BytesIO(xls_bytes), data_only=True)
        logger.info(f"Excel abierto con {len(wb.sheetnames)} hojas: {wb.sheetnames[:5]}...")
        
        anio, mes_num = map(int, mes.split('-'))
        resultado_por_fecha: Dict[date, Tuple[Dict[str, Decimal], Dict[str, bool]]] = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Intentar extraer día del nombre de la hoja (ej: "1", "01", "Sep-01", etc.)
            dia_num = None
            try:
                # Intenta parsear el nombre directamente como número
                dia_num = int(sheet_name)
            except:
                # Buscar dígitos en el nombre
                import re
                match = re.search(r'\d+', sheet_name)
                if match:
                    dia_num = int(match.group())
            
            if not dia_num or dia_num > 31:
                logger.warning(f"No se pudo determinar día de la hoja '{sheet_name}', saltando")
                continue
            
            try:
                fecha_hoja = date(anio, mes_num, dia_num)
            except ValueError:
                logger.warning(f"Fecha inválida: {anio}-{mes_num}-{dia_num}, saltando hoja '{sheet_name}'")
                continue
            
            logger.info(f"Procesando hoja '{sheet_name}' -> fecha {fecha_hoja}")
            
            # Parsear esta hoja
            try:
                mapping, usd_map = ImportadorSaldosService._parse_single_sheet(ws, etiqueta_fila_objetivo, logger)
                if mapping:
                    resultado_por_fecha[fecha_hoja] = (mapping, usd_map)
                    logger.info(f"✓ Hoja '{sheet_name}': {len(mapping)} cuentas encontradas")
            except Exception as e:
                logger.error(f"✗ Error en hoja '{sheet_name}': {str(e)}")
        
        return resultado_por_fecha
    
    @staticmethod
    def _parse_single_sheet(ws, etiqueta_fila_objetivo: str, logger) -> Tuple[Dict[str, Decimal], Dict[str, bool]]:
        """Parsea una sola hoja buscando la fila del concepto objetivo"""
        
        # Encontrar fila con etiqueta objetivo
        fila_objetivo = None
        fila_idx = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for cell in row:
                if isinstance(cell, str) and cell.strip().upper() == etiqueta_fila_objetivo:
                    fila_objetivo = row
                    fila_idx = idx
                    break
            if fila_objetivo:
                break
        
        if not fila_objetivo or fila_idx is None:
            raise ValueError(f"No se encontró '{etiqueta_fila_objetivo}' en esta hoja")

        # Buscar fila de números de cuenta (encabezado) arriba de fila objetivo
        numeros_cuenta = []
        fila_encabezado_idx = None
        patron = re.compile(r"\b\d{6,}\b")
        
        for offset in range(1, 5):  # buscar hasta 4 filas arriba
            if fila_idx - offset < 1:
                break
            fila_arriba = [c for c in ws.iter_rows(min_row=fila_idx-offset, max_row=fila_idx-offset, values_only=True)][0]
            encontrados = []
            for c in fila_arriba:
                if isinstance(c, (int, float)):
                    encontrados.append(str(int(c)))
                elif isinstance(c, str) and patron.search(c):
                    encontrados.append(patron.search(c).group(0))
                else:
                    encontrados.append(None)
            
            if len([x for x in encontrados if x]) >= 2:
                numeros_cuenta = encontrados
                fila_encabezado_idx = fila_idx - offset
                break
        
        if not numeros_cuenta:
            raise ValueError("No se encontró fila con números de cuenta")

        # Detectar columnas USD
        columnas_usd = [False] * len(numeros_cuenta)
        if fila_encabezado_idx:
            for offset_moneda in range(1, 4):
                if fila_encabezado_idx - offset_moneda < 1:
                    break
                fila_moneda = [c for c in ws.iter_rows(min_row=fila_encabezado_idx-offset_moneda, max_row=fila_encabezado_idx-offset_moneda, values_only=True)][0]
                for idx_col, cell in enumerate(fila_moneda):
                    if idx_col < len(columnas_usd) and isinstance(cell, str):
                        cell_upper = cell.strip().upper()
                        if 'USD' in cell_upper or 'DOLAR' in cell_upper or 'DÓLAR' in cell_upper or 'US$' in cell_upper:
                            columnas_usd[idx_col] = True

        # Leer valores de la fila objetivo
        valores = []
        for c in fila_objetivo:
            if isinstance(c, (int, float)):
                valores.append(Decimal(str(c)))
            elif isinstance(c, str):
                cleaned = c.replace(',', '').replace('(', '-').replace(')', '').strip()
                try:
                    valores.append(Decimal(cleaned))
                except Exception:
                    valores.append(Decimal('0'))
            else:
                valores.append(Decimal('0'))

        # Emparejar por posición
        mapping: Dict[str, Decimal] = {}
        usd_map: Dict[str, bool] = {}
        for i, cuenta in enumerate(numeros_cuenta):
            if cuenta and i < len(valores):
                mapping[cuenta] = valores[i]
                usd_map[cuenta] = columnas_usd[i] if i < len(columnas_usd) else False
        
        return mapping, usd_map

    @staticmethod
    def _obtener_trm(db: Session, fecha: date) -> Optional[Decimal]:
        trm = db.query(TRM).filter(TRM.fecha == fecha).first()
        return Decimal(str(trm.valor)) if trm else None

    @staticmethod
    def importar(
        db: Session,
        tipo_carga: str,
        mes: str,
        dia: Optional[str],
        sobrescribir: bool,
        archivo_excel: bytes,
        usuario_id: int = 1
    ) -> Dict:
        resultado = ImportadorSaldosResult()
        resultado.tipo_carga = tipo_carga
        resultado.mes = mes
        resultado.dia = dia
        resultado.overwrite = sobrescribir

        # Validar mes
        try:
            anio, mes_num = map(int, mes.split('-'))
            primer_dia_mes = date(anio, mes_num, 1)
        except Exception:
            raise ValueError("Formato de mes inválido. Use YYYY-MM")

        hoy = date.today()
        if tipo_carga == 'dia':
            if not dia:
                raise ValueError("Debe enviar dia cuando tipo_carga=dia")
            fecha_dia = datetime.strptime(dia, '%Y-%m-%d').date()
            if fecha_dia >= hoy:
                raise ValueError("Solo se pueden cargar días anteriores al vigente")
            dias_a_procesar = [fecha_dia]
        else:
            # mes completo hasta día anterior
            dias_a_procesar = []
            cursor = primer_dia_mes
            while cursor < hoy:
                dias_a_procesar.append(cursor)
                cursor += timedelta(days=1)

        # Parsear Excel (multi-hoja) - mismo archivo para ambos conceptos
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            datos_por_fecha = ImportadorSaldosService._parse_excel_multi_sheet(
                archivo_excel, 'SALDO INICIAL', mes
            )
            logger.info(f"✓ Excel parseado: {len(datos_por_fecha)} días encontrados")
        except Exception as e:
            resultado.errores.append(f"Error parseando Excel: {str(e)}")
            datos_por_fecha = {}

        # Obtener conceptos
        concepto_tes = db.query(ConceptoFlujoCaja).filter(
            ConceptoFlujoCaja.nombre == 'SALDO INICIAL', 
            ConceptoFlujoCaja.area == 'tesoreria'
        ).first()
        concepto_pag = db.query(ConceptoFlujoCaja).filter(
            ConceptoFlujoCaja.nombre == 'SALDO DIA ANTERIOR', 
            ConceptoFlujoCaja.area == 'pagaduria'
        ).first()
        
        if not concepto_tes:
            raise ValueError("Concepto 'SALDO INICIAL' (tesoreria) no existe")
        if not concepto_pag:
            raise ValueError("Concepto 'SALDO DIA ANTERIOR' (pagaduria) no existe")

        # Mapeo numero cuenta -> cuenta_bancaria
        cuentas_bancarias = db.query(CuentaBancaria).all()
        cuentas_map: Dict[str, CuentaBancaria] = {c.numero_cuenta: c for c in cuentas_bancarias if c.numero_cuenta}
        logger.info(f"Total cuentas en BD: {len(cuentas_map)}")

        for dia_trabajo in dias_a_procesar:
            trm_valor = ImportadorSaldosService._obtener_trm(db, dia_trabajo)
            if not trm_valor:
                resultado.dias_sin_trm.append(dia_trabajo.isoformat())
                continue
            
            # Verificar si hay datos para este día
            if dia_trabajo not in datos_por_fecha:
                continue
            
            mapa_valores, usd_flags = datos_por_fecha[dia_trabajo]
            logger.info(f"Día {dia_trabajo}: procesando {len(mapa_valores)} cuentas")
            
            for numero_cuenta, valor in mapa_valores.items():
                cuenta = cuentas_map.get(numero_cuenta)
                if not cuenta:
                    if numero_cuenta not in resultado.cuentas_sin_match:
                        resultado.cuentas_sin_match.append(numero_cuenta)
                    continue
                
                es_usd = usd_flags.get(numero_cuenta, False)
                monto_insertar = valor * trm_valor / 1000 if es_usd else valor
                
                # Crear/Actualizar TESORERÍA
                existente_tes = db.query(TransaccionFlujoCaja).filter(
                    TransaccionFlujoCaja.fecha == dia_trabajo,
                    TransaccionFlujoCaja.concepto_id == concepto_tes.id,
                    TransaccionFlujoCaja.cuenta_id == cuenta.id,
                    TransaccionFlujoCaja.area == AreaTransaccion.tesoreria
                ).first()
                
                if existente_tes and sobrescribir:
                    existente_tes.monto = monto_insertar
                    existente_tes.descripcion = 'Importación Excel sobrescrita'
                    resultado.cuentas_tesoreria += 1
                elif not existente_tes:
                    db.add(TransaccionFlujoCaja(
                        concepto_id=concepto_tes.id,
                        cuenta_id=cuenta.id,
                        compania_id=cuenta.compania_id,
                        fecha=dia_trabajo,
                        monto=monto_insertar,
                        descripcion='Importación Excel',
                        usuario_id=usuario_id,
                        area=AreaTransaccion.tesoreria
                    ))
                    resultado.cuentas_tesoreria += 1
                
                # Crear/Actualizar PAGADURÍA (mismo valor)
                existente_pag = db.query(TransaccionFlujoCaja).filter(
                    TransaccionFlujoCaja.fecha == dia_trabajo,
                    TransaccionFlujoCaja.concepto_id == concepto_pag.id,
                    TransaccionFlujoCaja.cuenta_id == cuenta.id,
                    TransaccionFlujoCaja.area == AreaTransaccion.pagaduria
                ).first()
                
                if existente_pag and sobrescribir:
                    existente_pag.monto = monto_insertar
                    existente_pag.descripcion = 'Importación Excel sobrescrita'
                    resultado.cuentas_pagaduria += 1
                elif not existente_pag:
                    db.add(TransaccionFlujoCaja(
                        concepto_id=concepto_pag.id,
                        cuenta_id=cuenta.id,
                        compania_id=cuenta.compania_id,
                        fecha=dia_trabajo,
                        monto=monto_insertar,
                        descripcion='Importación Excel',
                        usuario_id=usuario_id,
                        area=AreaTransaccion.pagaduria
                    ))
                    resultado.cuentas_pagaduria += 1
            
            resultado.dias_procesados += 1
        
        # Deduplicar cuentas_sin_match
        resultado.cuentas_sin_match = sorted(set(resultado.cuentas_sin_match))
        resultado.cuentas_procesadas = resultado.cuentas_tesoreria + resultado.cuentas_pagaduria
        
        db.commit()
        return resultado.to_dict()
