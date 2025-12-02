"""
Script para diagnosticar diferencias entre cuentas del Excel y la BD
"""
from openpyxl import load_workbook
from app.core.database import engine
from sqlalchemy import text
import re

# Ruta al Excel (ajustar según tu archivo)
EXCEL_PATH = input("Ingresa la ruta completa al archivo Excel: ").strip('"')

print(f"\n{'='*80}")
print("ANÁLISIS DE CUENTAS - EXCEL vs BASE DE DATOS")
print(f"{'='*80}\n")

# 1. Extraer cuentas del Excel
print("1. EXTRAYENDO CUENTAS DEL EXCEL...")
print("-" * 80)

try:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    # Tomar la primera hoja como muestra
    ws = wb[wb.sheetnames[0]]
    
    # Buscar la fila "SALDO INICIAL"
    target_row_idx = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50), start=1):
        cell_val = str(row[0].value or '').strip().upper()
        if 'SALDO INICIAL' in cell_val:
            target_row_idx = idx
            print(f"✓ Encontrada fila 'SALDO INICIAL' en fila {idx}")
            break
    
    if not target_row_idx:
        print("✗ No se encontró la fila 'SALDO INICIAL'")
        exit(1)
    
    # Extraer números de cuenta del encabezado (fila anterior)
    header_row = list(ws.iter_rows(min_row=target_row_idx - 1, max_row=target_row_idx - 1))[0]
    
    cuentas_excel = []
    cuenta_pattern = re.compile(r'\d{6,}')
    
    for idx, cell in enumerate(header_row):
        cell_value = str(cell.value or '').strip()
        match = cuenta_pattern.search(cell_value)
        if match:
            numero_cuenta = match.group()
            cuentas_excel.append({
                'numero': numero_cuenta,
                'columna': idx,
                'header_completo': cell_value
            })
    
    print(f"✓ Encontradas {len(cuentas_excel)} cuentas en el Excel\n")
    
    print("CUENTAS ENCONTRADAS EN EXCEL:")
    print("-" * 80)
    for c in cuentas_excel[:10]:  # Mostrar solo primeras 10
        print(f"  Columna {c['columna']:3d} | {c['numero']:15s} | {c['header_completo']}")
    if len(cuentas_excel) > 10:
        print(f"  ... y {len(cuentas_excel) - 10} cuentas más")
    
    wb.close()
    
except Exception as e:
    print(f"✗ Error leyendo Excel: {e}")
    exit(1)

# 2. Extraer cuentas de la BD
print(f"\n2. EXTRAYENDO CUENTAS DE LA BASE DE DATOS...")
print("-" * 80)

try:
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT id, numero_cuenta, nombre, compania_id, activa
            FROM cuentas_bancarias
            ORDER BY numero_cuenta
        """))
        cuentas_bd = result.fetchall()
        
    print(f"✓ Encontradas {len(cuentas_bd)} cuentas en la BD\n")
    
    print("PRIMERAS 10 CUENTAS EN BD:")
    print("-" * 80)
    for c in cuentas_bd[:10]:
        print(f"  ID: {c[0]:3d} | Número: {c[1]:15s} | Nombre: {c[2][:40]:40s} | Activa: {c[4]}")
    if len(cuentas_bd) > 10:
        print(f"  ... y {len(cuentas_bd) - 10} cuentas más")
    
except Exception as e:
    print(f"✗ Error leyendo BD: {e}")
    exit(1)

# 3. Comparar y encontrar diferencias
print(f"\n3. ANÁLISIS DE COINCIDENCIAS")
print("=" * 80)

numeros_bd = {str(c[1]).strip() for c in cuentas_bd if c[1]}
numeros_excel = {c['numero'] for c in cuentas_excel}

# Cuentas en Excel pero NO en BD
no_encontradas = numeros_excel - numeros_bd
encontradas = numeros_excel & numeros_bd

print(f"\n✓ CUENTAS QUE SÍ COINCIDEN: {len(encontradas)}/{len(numeros_excel)}")
if encontradas:
    print("  Ejemplos:")
    for num in list(encontradas)[:5]:
        print(f"    - {num}")

print(f"\n✗ CUENTAS EN EXCEL PERO NO EN BD: {len(no_encontradas)}")
if no_encontradas:
    print("  ESTAS CUENTAS NO SE PODRÁN IMPORTAR:")
    for num in sorted(no_encontradas):
        # Buscar el header completo
        cuenta_info = next((c for c in cuentas_excel if c['numero'] == num), None)
        if cuenta_info:
            print(f"    - {num:15s} | {cuenta_info['header_completo']}")

# Cuentas en BD pero NO en Excel (solo informativo)
en_bd_no_excel = numeros_bd - numeros_excel
print(f"\nℹ CUENTAS EN BD PERO NO EN EXCEL: {len(en_bd_no_excel)} (esto es normal)")

# 4. Recomendaciones
print(f"\n{'='*80}")
print("RECOMENDACIONES")
print("=" * 80)

if no_encontradas:
    print("\n⚠ ACCIÓN REQUERIDA:")
    print("  Debes agregar estas cuentas a la BD antes de importar:")
    print("\n  Opción 1: Agregar manualmente desde el frontend")
    print("  Opción 2: Crear script SQL para insertar las cuentas faltantes")
    print("\n  Ejemplo SQL:")
    for num in list(no_encontradas)[:3]:
        cuenta_info = next((c for c in cuentas_excel if c['numero'] == num), None)
        if cuenta_info:
            nombre = cuenta_info['header_completo'].replace(num, '').strip()
            print(f"  INSERT INTO cuentas_bancarias (numero_cuenta, nombre, compania_id, activa) ")
            print(f"    VALUES ('{num}', '{nombre}', 1, 1);")
else:
    print("\n✓ TODAS LAS CUENTAS DEL EXCEL EXISTEN EN LA BD")
    print("  La importación debería funcionar correctamente.")

print(f"\n{'='*80}\n")
